"""
Convert Cisco Firepower Access Control Policy PDF reports to Excel.

Each rule becomes one row; each field becomes a column.
Multiple PDF files in the same directory are exported as separate sheets.
"""

import re
import sys
import subprocess
import threading
import time
from pathlib import Path


REQUIRED_PACKAGES = {
    "pdfminer": "pdfminer.six",
    "openpyxl": "openpyxl",
}


def ensure_dependencies():
    missing = []
    for import_name, pip_name in REQUIRED_PACKAGES.items():
        try:
            __import__(import_name)
        except ImportError:
            missing.append(pip_name)

    if not missing:
        return

    print(f"Missing libraries: {', '.join(missing)}")
    print("Installing...")
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])
        print("Installation complete.\n")
    except subprocess.CalledProcessError:
        print("ERROR: Failed to install required libraries.")
        print(f"Please run manually: pip install {' '.join(missing)}")
        sys.exit(1)


ensure_dependencies()

from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextBox
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Enable ANSI escape codes on Windows
if sys.platform == "win32":
    import os
    os.system("")

_BAR_WIDTH = 40
_GREEN  = "\033[32m"
_YELLOW = "\033[33m"
_RED    = "\033[31m"
_RESET  = "\033[0m"


class LoadingBar:
    """
    Context manager: shows an animated elapsed-time bar while a block runs.

    Usage:
        with LoadingBar("Parsing foo.pdf"):
            rules = parse_pdf(path)
    """

    def __init__(self, label: str = "", ref_seconds: float = 60.0):
        self._label       = label
        self._ref         = ref_seconds   # seconds at which bar reaches 100%
        self._stop        = threading.Event()
        self._ready       = threading.Event()  # signals initial line is printed
        self._thread      = None
        self._start_time  = None

    def _color(self, ratio: float) -> str:
        if ratio < 0.5:
            return _GREEN
        if ratio < 0.8:
            return _YELLOW
        return _RED

    def _render(self, elapsed: float) -> str:
        ratio  = min(elapsed / self._ref, 1.0)
        filled = int(_BAR_WIDTH * ratio)
        bar    = "█" * filled + "░" * (_BAR_WIDTH - filled)
        hms    = time.strftime("%H:%M:%S", time.gmtime(elapsed))
        color  = self._color(ratio)
        return f"  {color}[{bar}]{_RESET} {hms}"

    def _run(self):
        # Print initial blank bar on its own line, then signal ready
        sys.stdout.write(f"  [{'░' * _BAR_WIDTH}] 00:00:00\n")
        sys.stdout.flush()
        self._ready.set()

        while not self._stop.wait(timeout=1.0):
            elapsed = time.time() - self._start_time
            line = self._render(elapsed)
            sys.stdout.write(f"\033[1A\r\033[2K{line}\n")
            sys.stdout.flush()

    def __enter__(self):
        self._start_time = time.time()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()
        self._ready.wait()  # ensure bar line exists before caller runs
        return self

    def __exit__(self, *_):
        self._stop.set()
        self._thread.join()
        elapsed = time.time() - self._start_time
        bar  = "█" * _BAR_WIDTH
        hms  = time.strftime("%H:%M:%S", time.gmtime(elapsed))
        line = f"  {_GREEN}[{bar}]{_RESET} {hms}  {_GREEN}✓{_RESET}"
        sys.stdout.write(f"\033[1A\r\033[2K{line}\n")
        sys.stdout.flush()


FIELD_NAMES = [
    "Action",
    "Source Zones",
    "Destination Zones",
    "Source Tunnels",
    "Source Networks",
    "Original Client Networks",
    "Destination Networks",
    "Safe Search",
    "Youtube EDU",
    "VLAN Tags",
    "Users",
    "Application Filters",
    "Source Ports",
    "Destination Ports",
    "Source ISE Metadata",
    "Destination ISE Metadata",
    "Security Group Tag",
    "Time Range",
    "URLs",
    "Intrusion Policy",
    "Variable Set",
    "File Policy",
    "Log at Beginning of Connection",
    "Log at End of Connection",
    "Log File Events",
    "Send Events to Defense Center",
    "Send using specific syslog alert",
    "Comments",
]

FIELD_SET = set(FIELD_NAMES)

# Columns that appear in the Excel output before field columns
META_COLUMNS = ["Policy", "Rule Number", "Rule Name", "Enabled", "Section"]

ALL_COLUMNS = META_COLUMNS + FIELD_NAMES

# x-coordinate threshold separating label column from value column
LABEL_X_MAX = 200

# Page number elements are single digits/short numbers centered; skip them
PAGE_NUM_PATTERN = re.compile(r"^\d{1,3}$")

# Rule header: "1:RuleName" or "1:RuleName(disable)"
RULE_HEADER_PATTERN = re.compile(r"^(\d+):(.+)$")


def extract_page_elements(page_layout):
    """Return sorted list of (x0, y0, y1, text) from a pdfminer page."""
    elements = []
    for element in page_layout:
        if not isinstance(element, LTTextBox):
            continue
        text = element.get_text().strip()
        if not text:
            continue
        # Skip standalone page numbers
        if PAGE_NUM_PATTERN.match(text):
            continue
        elements.append((element.x0, element.y0, element.y1, text))
    return elements


def find_value_for_label(label_y0, right_elements, tolerance=5):
    """
    Match a label to its value box.
    In a two-column PDF the label's y0 falls within the value box's
    vertical span [y0_val, y1_val] (multi-line values extend downward).
    """
    for _, vy0, vy1, vtext in right_elements:
        if vy0 - tolerance <= label_y0 <= vy1 + tolerance:
            return vtext
    return ""


def parse_pdf(pdf_path: Path):
    """
    Parse a Firepower policy PDF and return a list of rule dicts.
    Keys: META_COLUMNS + FIELD_NAMES
    """
    policy_name = pdf_path.stem
    rules = []
    current_rule = None
    current_section = "Mandatory Rules"

    for page_layout in extract_pages(str(pdf_path)):
        elements = extract_page_elements(page_layout)

        # Split into left (labels/headers) and right (values) columns
        left_elements = [(x0, y0, y1, t) for x0, y0, y1, t in elements if x0 < LABEL_X_MAX]
        right_elements = [(x0, y0, y1, t) for x0, y0, y1, t in elements if x0 >= LABEL_X_MAX]

        # Sort left elements top-to-bottom (y1 desc)
        left_elements.sort(key=lambda e: -e[2])

        for x0, y0, y1, text in left_elements:
            # Track section boundaries
            if text in ("Mandatory Rules", "Default Rules"):
                current_section = text
                continue

            # Rule header detection
            m = RULE_HEADER_PATTERN.match(text)
            if m and x0 < 50:
                rule_num = m.group(1)
                full_name = m.group(2)
                # Parse "(disable)" / "(enable)" suffix
                enabled = "Yes"
                name_clean = full_name
                suffix_m = re.search(r"\((disable|enable)\)$", full_name, re.IGNORECASE)
                if suffix_m:
                    enabled = "No" if suffix_m.group(1).lower() == "disable" else "Yes"
                    name_clean = full_name[: suffix_m.start()].strip()

                current_rule = {col: "" for col in ALL_COLUMNS}
                current_rule["Policy"] = policy_name
                current_rule["Rule Number"] = rule_num
                current_rule["Rule Name"] = name_clean
                current_rule["Enabled"] = enabled
                current_rule["Section"] = current_section
                rules.append(current_rule)
                continue

            # Known field label
            if text in FIELD_SET and current_rule is not None:
                value = find_value_for_label(y0, right_elements)
                # Normalise multi-line values: join with newline preserved
                current_rule[text] = value
                continue

    return rules


def apply_header_style(ws, header_row=1):
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx in range(1, len(ALL_COLUMNS) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def apply_data_style(ws, num_rows, action_col_idx):
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    action_colors = {
        "allow": "E2EFDA",
        "block": "FCE4D6",
        "monitor": "FFF2CC",
        "trust": "DAEEF3",
    }

    for row_idx in range(2, num_rows + 2):
        action_cell = ws.cell(row=row_idx, column=action_col_idx)
        action_val = str(action_cell.value or "").lower()
        row_fill = None
        for key, color in action_colors.items():
            if key in action_val:
                row_fill = PatternFill("solid", fgColor=color)
                break

        for col_idx in range(1, len(ALL_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if row_fill:
                cell.fill = row_fill


def set_column_widths(ws):
    width_map = {
        "Policy": 20,
        "Rule Number": 8,
        "Rule Name": 30,
        "Enabled": 8,
        "Section": 16,
        "Action": 10,
        "Source Zones": 20,
        "Destination Zones": 20,
        "Source Tunnels": 15,
        "Source Networks": 35,
        "Original Client Networks": 35,
        "Destination Networks": 35,
        "Safe Search": 10,
        "Youtube EDU": 10,
        "VLAN Tags": 12,
        "Users": 20,
        "Application Filters": 20,
        "Source Ports": 20,
        "Destination Ports": 20,
        "Source ISE Metadata": 20,
        "Destination ISE Metadata": 20,
        "Security Group Tag": 15,
        "Time Range": 15,
        "URLs": 25,
        "Intrusion Policy": 30,
        "Variable Set": 15,
        "File Policy": 20,
        "Log at Beginning of Connection": 12,
        "Log at End of Connection": 12,
        "Log File Events": 12,
        "Send Events to Defense Center": 12,
        "Send using specific syslog alert": 12,
        "Comments": 30,
    }
    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_map.get(col_name, 15)


def write_sheet(wb, sheet_name, rules):
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name max 31 chars

    # Header row
    for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, rule in enumerate(rules, start=2):
        for col_idx, col_name in enumerate(ALL_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=rule.get(col_name, ""))

    # Freeze top row
    ws.freeze_panes = "A2"

    # Styling
    apply_header_style(ws)
    action_col = ALL_COLUMNS.index("Action") + 1
    apply_data_style(ws, len(rules), action_col)
    set_column_widths(ws)

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_COLUMNS))}1"

    return ws


def main():
    script_dir = Path(__file__).parent
    input_dir = script_dir / "input"
    output_dir = script_dir / "output"

    if not input_dir.exists():
        input_dir.mkdir()
        print(f"Created input folder: {input_dir}")
        print("Place your PDF files in the 'input' folder and run again.")
        sys.exit(0)

    output_dir.mkdir(exist_ok=True)

    pdf_files = sorted(input_dir.glob("*.pdf"))

    if not pdf_files:
        print(f"No PDF files found in: {input_dir}")
        sys.exit(1)

    print(f"Found {len(pdf_files)} PDF file(s):")
    for p in pdf_files:
        print(f"  {p.name}")

    total_rules = 0
    for pdf_path in pdf_files:
        print(f"\nParsing: {pdf_path.name}")
        with LoadingBar(pdf_path.name):
            rules = parse_pdf(pdf_path)
        print(f"  -> {len(rules)} rules extracted")
        total_rules += len(rules)

        wb = Workbook()
        wb.remove(wb.active)
        write_sheet(wb, pdf_path.stem, rules)

        output_path = output_dir / f"{pdf_path.stem}.xlsx"
        wb.save(str(output_path))
        print(f"     Saved: output/{pdf_path.stem}.xlsx")

    print(f"\nDone. {total_rules} total rules written to: {output_dir}")


if __name__ == "__main__":
    main()
